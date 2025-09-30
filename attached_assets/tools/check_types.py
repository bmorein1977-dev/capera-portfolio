from openpyxl import load_workbook
import sys

def norm(s):
    if s is None: return ""
    t = str(s).strip().lower().replace("-", " ").replace("_"," ")
    return " ".join(t.split())

def guess_type(t):
    if not t: return ""
    if t in {"knowledge","k","uk","u/k","underpinning","under pinning","underpinning knowledge","underpin"}: return "knowledge"
    if t.startswith("underpin") or t.startswith("knowledg"): return "knowledge"
    if t in {"performance","p","pc","p/c","perf","performance criteria","performance criterion"}: return "performance"
    if t.startswith("perform"): return "performance"
    return f"? ({t})"

if len(sys.argv)<2:
    print("Usage: python tools/check_types.py <xlsx>")
    sys.exit(1)

wb = load_workbook(sys.argv[1])
ws = wb.active

hdr = [ (i+1, (c.value if c else "")) for i,c in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=False))) ]
print("Headers:", hdr)
idx = None
for i,c in hdr:
    h = norm(c)
    if h in {"criteria","criteria type","type"}:
        idx = i; break
if not idx:
    print("No Type/Criteria column found"); sys.exit(2)

unknown = 0
for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    raw = row[idx-1]
    t = norm(raw)
    g = guess_type(t)
    if g.startswith("?"):
        print(f"Row {ridx}: Unrecognised type '{raw}'")
        unknown += 1
print(f"Unknown count: {unknown}")
