
from openpyxl import load_workbook
import sys

HEADER_ALIASES = {
  "competence category":"category","competency category":"category","category":"category",
  "competence element":"element","competency element":"element","element":"element",
  "proficiency level":"prof_level","proficiency":"prof_level",
  "criteria":"criteria_type","subcategory":"subcategory",
  "assessment criteria":"criteria_text","criteria text":"criteria_text",
  "assessor guidance":"guidance","criticality rating":"criticality",
  "reassessment validity":"reassess_years","required":"required_flag",
}
def norm(s): return (str(s).strip().lower() if s is not None else "")

if len(sys.argv) < 2:
    print("Usage: python tools/validate_xlsx.py <path.xlsx>")
    sys.exit(1)

path = sys.argv[1]
wb = load_workbook(path)
ws = wb.active
hdr_cells = list(ws.iter_rows(min_row=1,max_row=1,values_only=False))[0]
mapped = {}
raw = []
for i,c in enumerate(hdr_cells,1):
    k = HEADER_ALIASES.get(norm(c.value))
    if k: mapped[k] = i
    raw.append((i, c.value))

print("Raw headers:", raw)
print("Mapped columns:", mapped)
need = ["category","element","criteria_type","subcategory","criteria_text"]
missing=[k for k in need if k not in mapped]
print("Missing:", missing)
