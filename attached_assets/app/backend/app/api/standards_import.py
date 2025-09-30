
from fastapi import APIRouter, UploadFile, File, Query, Depends, HTTPException
from sqlalchemy.orm import Session
from io import BytesIO
from typing import Dict, Any, List, Tuple
from app.db import get_db
from app.dependencies_roles import require_roles

router = APIRouter()

HEADER_ALIASES = {
    "competence category": "category",
    "competency category": "category",
    "category": "category",
    "competence element": "element",
    "competency element": "element",
    "element": "element",
    "proficiency level": "prof_level",
    "proficiency": "prof_level",
    "criteria": "criteria_type",
    "subcategory": "subcategory",
    "assessment criteria": "criteria_text",
    "criteria text": "criteria_text",
    "assessor guidance": "guidance",
    "criticality rating": "criticality",
    "reassessment validity": "reassess_years",
    "required": "required_flag",
}

def norm_header(h: Any) -> str:
    if h is None: return ""
    return str(h).strip().lower()

def norm_value(v: Any) -> Any:
    if isinstance(v, str):
        s = v.strip()
        return s if s != "" else None
    return v

def norm_criteria_type(s: Any) -> str:
    if not s: return ""
    s = str(s).strip().lower()
    if s.startswith("underpinning"): return "knowledge"
    if s.startswith("performance"):  return "performance"
    return s

def is_mandatory(flag: Any) -> bool:
    return str(flag).strip().upper() == "M" if flag is not None else False

def fmt_bool(cell) -> tuple[bool,bool]:
    try:
        f = cell.font
        return bool(getattr(f, "bold", False)), bool(getattr(f, "italic", False))
    except Exception:
        return (False, False)

@router.post("/import-xlsx")
async def import_xlsx(
    file: UploadFile = File(...),
    dry_run: bool = Query(True),
    client_id: int | None = None,
    db: Session = Depends(get_db),
    user = Depends(require_roles("Admin","Developer"))
):
    name = (file.filename or "").lower()
    if not name.endswith(".xlsx"):
        raise HTTPException(400, "Please upload an .xlsx file")

    content = await file.read()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(content), data_only=True, read_only=False)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"openpyxl failed to read workbook: {e}")

    header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
    columns: Dict[str,int] = {}
    for idx, cell in enumerate(header_cells, start=1):
        key = HEADER_ALIASES.get(norm_header(cell.value), None)
        if key:
            columns[key] = idx

    raw_headers = [ (i+1, (c.value if c and c.value is not None else "")) for i, c in enumerate(header_cells) ]
    print("IMPORT HEADERS ==>", raw_headers)

    required_cols = ["category","element","criteria_type","subcategory","criteria_text"]
    missing = [c for c in required_cols if c not in columns]

    if missing:
        values = [norm_header(c.value) for c in header_cells]
        if "competence element" in values: columns["element"] = values.index("competence element")+1
        if "subcategory" in values: columns["subcategory"] = values.index("subcategory")+1
        missing = [c for c in required_cols if c not in columns]

    if missing:
        raise HTTPException(400, f"Missing columns: {missing}. Found {columns}. "
                                 f"Tip: headers may contain trailing spaces; the importer trims and maps them.")

    plan: Dict[str, Any] = {}

    current = {
        "category": None,
        "element": None,
        "prof_level": None,
        "criteria_type": None,
        "subcategory": None,
        "criticality": None,
        "reassess_years": None,
    }

    for r in ws.iter_rows(min_row=2, values_only=False):
        def col(k): 
            idx = columns.get(k)
            return r[idx-1] if idx else None

        cat_c     = col("category")
        elem_c    = col("element")
        prof_c    = col("prof_level")
        crittype_c= col("criteria_type")
        subcat_c  = col("subcategory")
        text_c    = col("criteria_text")
        guid_c    = col("guidance") if "guidance" in columns else None
        critrat_c = col("criticality") if "criticality" in columns else None
        reass_c   = col("reassess_years") if "reassess_years" in columns else None
        req_c     = col("required_flag") if "required_flag" in columns else None

        values = [cell.value if cell else None for cell in [cat_c, elem_c, crittype_c, subcat_c, text_c, guid_c, critrat_c, reass_c, req_c]]
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in values):
            continue

        cat  = norm_value(cat_c.value)  if cat_c else None
        elem = norm_value(elem_c.value) if elem_c else None
        current["category"]      = cat  or current["category"]
        current["element"]       = elem or current["element"]
        current["prof_level"]    = norm_value(prof_c.value) if (prof_c and prof_c.value is not None) else current["prof_level"]
        current["criteria_type"] = norm_criteria_type(crittype_c.value if crittype_c else current["criteria_type"])
        current["subcategory"]   = norm_value(subcat_c.value) if (subcat_c and subcat_c.value is not None) else current["subcategory"]
        current["criticality"]   = norm_value(critrat_c.value) if (critrat_c and critrat_c.value is not None) else current["criticality"]
        current["reassess_years"]= int(norm_value(reass_c.value)) if (reass_c and norm_value(reass_c.value) not in (None,"")) else current["reassess_years"]

        if not current["subcategory"]:
            current["subcategory"] = "General"

        text = norm_value(text_c.value if text_c else None)
        if not text:
            continue

        key = f"{current['category']}::{current['element']}"
        if key not in plan:
            plan[key] = {
                "category": current["category"],
                "element": current["element"],
                "proficiency_scheme": int(current["prof_level"]) if current["prof_level"] else 1,
                "criticality": current["criticality"] or "",
                "reassess_years": int(current["reassess_years"] or 0),
                "knowledge": [],
                "performance": [],
                "_counters": {
                    "knowledge": {"_order": [], "_map": {}},
                    "performance": {"_order": [], "_map": {}},
                }
            }

        entry = plan[key]
        ctype = current["criteria_type"] or "knowledge"
        if ctype not in ("knowledge","performance"):
            ctype = "knowledge"

        ctrs = entry["_counters"][ctype]
        sub = current["subcategory"]
        if sub not in ctrs["_map"]:
            ctrs["_order"].append(sub)
            ctrs["_map"][sub] = {"major": len(ctrs["_order"]), "minor": 1}
            entry[ctype].append({"title": sub, "items": []})

        major = ctrs["_map"][sub]["major"]
        minor = ctrs["_map"][sub]["minor"]
        ctrs["_map"][sub]["minor"] += 1

        prefix = "K" if ctype == "knowledge" else "P"
        num = f"{prefix} {major}.{minor}"

        t_bold, t_italic = fmt_bool(text_c)
        g_text  = norm_value(guid_c.value) if guid_c else None
        g_bold, g_italic = fmt_bool(guid_c) if guid_c else (False, False)
        gnum = f"{prefix}G {major}.{minor}" if g_text else None

        required = is_mandatory(req_c.value) if req_c else False

        section = next(sec for sec in entry[ctype] if sec["title"] == sub)
        section["items"].append({
            "number": num,
            "text": text,
            "fmt_bold": t_bold,
            "fmt_italic": t_italic,
            "assessor_guidance": g_text or "",
            "guidance_number": gnum,
            "guidance_fmt_bold": g_bold,
            "guidance_fmt_italic": g_italic,
            "required": required,
        })

    for e in plan.values():
        e.pop("_counters", None)

    if dry_run:
        return {"status":"ok","mode":"dry_run","standards": plan, "count": len(plan)}

    return {"status":"ok","mode":"commit_skipped","reason":"Wire commit logic to your ORM (see README).","standards": plan, "count": len(plan)}
